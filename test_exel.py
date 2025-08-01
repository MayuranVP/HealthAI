import streamlit as st
import speech_recognition as sr
from gtts import gTTS
import pygame
import tempfile
import time
import pandas as pd
from datetime import datetime
import os
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide")




# Traductions multilingues
translations = {
    "fr": {
        "title": "🧠 Simulateur ACLS Vocal Interactif",
        "start": "▶️ Lancer la simulation",
        "stop": "🛑 Stopper la simulation",
        "speak_now": " Parlez maintenant (oui ou non)...",
        "said": " Vous avez dit :",
        "no_sound": "⏱ Aucun son détecté dans le temps imparti.",
        "not_understood": " Réponse non comprise.",
        "end_algo": " Fin de l'algorithme ACLS",
        "export": "📤 Exporter le rapport de simulation",
        "download": "📥 Télécharger le rapport Excel",
        "time_left": "⏳ Temps restant :",
        "time_over": "⏱️ Temps écoulé.",
        "stopped": "🚨 Simulation interrompue.",
        "yes": "oui",
        "no": "non",
        "rcp_start": "Début de la RCP. Connexion de l'oxygène et du défibrillateur",
        "shock1": " CHOC 1 délivré",
        "shock2": " CHOC 2",
        "shock3": " CHOC 3",
        "shock_conv": " CHOC sur rythme devenu choquable",
        "epi": " Épinéphrine administrée",
        "epi_alone": " Épinéphrine seule + RCP",
        "epi_now": " Épinéphrine IMMÉDIATE",
        "amio": " Amiodarone/Lidocaïne + causes réversibles",
        "rcp_cause": " RCP 2 min + Causes réversibles",
        "rcp_loop": "🔁 Continuer RCP / choc / médicaments",
        "rosc_reached": "✅ ROSC atteint → soins post-arrêt",
        "prompt_rhythm": "Le rythme est-il choquable ?",
        "prompt_rhythm2": "Le rythme est-il encore choquable ?",
        "prompt_rhythm3": "Encore choquable ?",
        "prompt_nonshock1": "Le rythme est-il devenu choquable ?",
        "prompt_nonshock2": "Le rythme est-il devenu choquable ?",
        "prompt_rosc": "Y a-t-il un retour de circulation spontanée ?"
    },
    "en": {
        "title": "🧠 Interactive ACLS Voice Simulator",
        "start": "▶️ Start Simulation",
        "stop": "🛑 Stop Simulation",
        "speak_now": "🎙️ Speak now (yes or no)...",
        "said": "🗣️ You said:",
        "no_sound": "⏱️ No sound detected in time.",
        "not_understood": "🤔 Response not understood.",
        "end_algo": "🏁 End of ACLS algorithm",
        "export": "📤 Export simulation report",
        "download": "📥 Download Excel report",
        "time_left": "⏳ Time left:",
        "time_over": "⏱️ Time's up.",
        "stopped": "🚨 Simulation stopped.",
        "yes": "yes",
        "no": "no",
        "rcp_start": "Start CPR. Connect oxygen and defibrillator",
        "shock1": " SHOCK #1 delivered",
        "shock2": " SHOCK #2",
        "shock3": " SHOCK #3",
        "shock_conv": " SHOCK on converted rhythm",
        "epi": " Epinephrine administered",
        "epi_alone": " Epinephrine only + CPR",
        "epi_now": " IMMEDIATE Epinephrine",
        "amio": " Amiodarone/Lidocaine + reversible causes",
        "rcp_cause": " CPR 2 min + Reversible causes",
        "rcp_loop": " Continue CPR / shock / meds",
        "rosc_reached": " ROSC achieved → post-cardiac care",
        "prompt_rhythm": "Is the rhythm shockable?",
        "prompt_rhythm2": "Is the rhythm still shockable?",
        "prompt_rhythm3": "Still shockable?",
        "prompt_nonshock1": "Has the rhythm become shockable?",
        "prompt_nonshock2": "Has the rhythm become shockable?",
        "prompt_rosc": "Is there a return of spontaneous circulation?"
    },
    "th": {
        "title": "🧠 โปรแกรมจำลอง ACLS แบบใช้เสียง",
        "start": "▶️ เริ่มการจำลอง",
        "stop": "🛑 หยุดการจำลอง",
        "speak_now": " พูดเลย (ใช่ หรือ ไม่)...",
        "said": " คุณพูดว่า:",
        "no_sound": "⏱ ไม่พบเสียงในเวลาที่กำหนด",
        "not_understood": " ไม่เข้าใจคำตอบ",
        "end_algo": " สิ้นสุดขั้นตอน ACLS",
        "export": "📤 ส่งออกรายงานการจำลอง",
        "download": "📥 ดาวน์โหลดรายงาน Excel",
        "time_left": " เวลาที่เหลือ:",
        "time_over": "⏱ หมดเวลา",
        "stopped": " หยุดการจำลอง",
        "yes": "ใช่",
        "no": "ไม่ใช่",
        "rcp_start": "เริ่มการกดหน้าอก เชื่อมต่อออกซิเจนและเครื่องช็อกไฟฟ้า",
        "shock1": " ช็อกครั้งที่ 1",
        "shock2": " ช็อกครั้งที่ 2",
        "shock3": " ช็อกครั้งที่ 3",
        "shock_conv": " ช็อกจังหวะที่กลับมา",
        "epi": " ให้ยาอิพิเนฟริน",
        "epi_alone": " ให้ยาอิพิเนฟริน + ทำ CPR",
        "epi_now": " ให้ยาอิพิเนฟรินทันที",
        "amio": " ยา Amiodarone/Lidocaine + หาสาเหตุที่แก้ไขได้",
        "rcp_cause": "🔄 ทำ CPR + หาสาเหตุที่แก้ไขได้",
        "rcp_loop": "🔁 ทำ CPR / ช็อก / ให้ยา ต่อเนื่อง",
        "rosc_reached": "✅ ROSC สำเร็จ → ดูแลหลังหัวใจหยุดเต้น",
        "prompt_rhythm": "จังหวะสามารถช็อกได้หรือไม่?",
        "prompt_rhythm2": "จังหวะยังสามารถช็อกได้อยู่หรือไม่?",
        "prompt_rhythm3": "ยังสามารถช็อกได้อีกหรือไม่?",
        "prompt_nonshock1": "จังหวะเปลี่ยนเป็นสามารถช็อกได้หรือไม่?",
        "prompt_nonshock2": "จังหวะเปลี่ยนเป็นสามารถช็อกได้หรือไม่?",
        "prompt_rosc": "มีการไหลเวียนเลือดกลับมาหรือไม่?"
    }
}

# Sélecteur de langue dans la sidebar
st.sidebar.title("🌐 Langue / Language / ภาษา")
lang = st.sidebar.selectbox("Choisissez la langue", options=["fr", "en", "th"], format_func=lambda l: {"fr": "Français 🇫🇷", "en": "English 🇬🇧", "th": "ไทย 🇹🇭"}[l])
T = translations[lang]

# Synthèse vocale avec gTTS + pygame
def speak(text):
    try:
        tts = gTTS(text=text, lang=lang)
        temp_path = os.path.join(tempfile.gettempdir(), "temp_audio.mp3")
        tts.save(temp_path)
        pygame.mixer.init()
        pygame.mixer.music.load(temp_path)
        pygame.mixer.music.play()
        while pygame.mixer.music.get_busy():
            time.sleep(0.1)
        pygame.mixer.music.unload()
        os.remove(temp_path)
    except Exception as e:
        st.warning(f"Erreur de synthèse vocale : {e}")

# Session state initialisation
if "event_log" not in st.session_state:
    st.session_state.event_log = []
if "arret_demande" not in st.session_state:
    st.session_state.arret_demande = False
if "procedure_terminee" not in st.session_state:
    st.session_state.procedure_terminee = False

def log_event(message):
    timestamp = datetime.now().strftime("%H:%M:%S")
    st.markdown(f"🕒 `{timestamp}` - {message}")
    speak(message)
    st.session_state.event_log.append({"Horodatage": timestamp, "Événement": message})

def show_gif(key):
    gif_path = f"./gif/{key}.gif"
    if os.path.exists(gif_path):
        st.image(gif_path, use_column_width=True)

def message_key_from_text(text):
    for key, val in T.items():
        if val.strip() in text.strip():
            return key
    return ""

def listen():
    recognizer = sr.Recognizer()
    recognizer.energy_threshold = 300
    recognizer.pause_threshold = 0.8
    recognizer.non_speaking_duration = 0.5
    with sr.Microphone() as source:
        st.write(T["speak_now"])
        try:
            audio = recognizer.listen(source, timeout=6, phrase_time_limit=5)
            response = recognizer.recognize_google(audio, language={"fr": "fr-FR", "en": "en-US", "th": "th-TH"}[lang])
            st.success(f"{T['said']} {response}")
            return response.lower()
        except sr.WaitTimeoutError:
            st.warning(T["no_sound"])
        except sr.UnknownValueError:
            st.warning(T["not_understood"])
        except sr.RequestError as e:
            st.error(f"Speech API error: {e}")
        return None

def ask_question(question_text_key, retry=3):
    log_event(T[question_text_key])
    for _ in range(retry):
        if st.session_state.arret_demande:
            log_event(T["stopped"])
            return -1
        response = listen()
        if response:
            if T["yes"] in response:
                return 1
            elif T["no"] in response:
                return 0
    log_event(T["not_understood"])
    return -1

def countdown(seconds, message, gif_key=None):
    container = st.container()
    if gif_key:
        col1, col2 = container.columns([1, 1])
        with col1:
            st.image(f"./gif/{gif_key}.gif", use_container_width=True)
        timer_placeholder = col2.empty()
    else:
        timer_placeholder = container.empty()

    for remaining in range(seconds, 0, -1):
        if st.session_state.arret_demande:
            timer_placeholder.warning(T["stopped"])
            return
        timer_placeholder.markdown(
            f"<h1 style='text-align: center; color: #FF4B4B;'>{message} {remaining} sec</h1>",
            unsafe_allow_html=True
        )
        time.sleep(1)
    timer_placeholder.success(T["time_over"])

def step_safe(message_key):
    if st.session_state.arret_demande:
        log_event(T["stopped"])
        return False
    log_event(T[message_key])
    return True

def export_log_to_excel(nom, prenom):

    df = pd.DataFrame(st.session_state.event_log)

    wb = Workbook()
    ws = wb.active
    ws.title = "Journal"

    ws["A1"] = f"Nom : {prenom} {nom}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    start_row = 3
    header_fill = PatternFill(start_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            if r_idx == start_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 🔧 Correction ici
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer



def run_acls():
    st.title("")  # Vide car on affiche le titre autrement

    col1, col2, col3 = st.columns([1, 2, 1])  # Centre les boutons
    with col2:
        st.markdown(f"<h1 style='text-align: center;'>{T['title']}</h1>", unsafe_allow_html=True)
        start_clicked = st.button(T["start"], key="btn_start", use_container_width=True)
        stop_clicked = st.button(T["stop"], key="btn_stop", use_container_width=True)

    if stop_clicked:
        st.session_state.arret_demande = True

    if start_clicked:
        st.session_state.event_log.clear()
        st.session_state.arret_demande = False
        st.session_state.procedure_terminee = False
        if not step_safe("rcp_start"):
            return
        st.session_state.event_log.clear()
        st.session_state.arret_demande = False
        st.session_state.procedure_terminee = False
        if not step_safe("rcp_start"):
            return
        shockable = ask_question("prompt_rhythm")
        if shockable == 1:
            if not step_safe("shock1"): return
            countdown(10, T["time_left"], gif_key="rcp_start")
            shockable = ask_question("prompt_rhythm2")
            if shockable == 1:
                if not step_safe("shock2"): return
                if not step_safe("epi"): return
                countdown(10, T["time_left"], gif_key="rcp_start")
                shockable = ask_question("prompt_rhythm3")
                if shockable == 1:
                    if not step_safe("shock3"): return
                    if not step_safe("amio"): return
                    countdown(10, T["time_left"], gif_key="rcp_start")
                else:
                    if not step_safe("rcp_cause"): return
                    countdown(10, T["time_left"], gif_key="rcp_start")
            else:
                if not step_safe("epi_alone"): return
                countdown(10, T["time_left"], gif_key="rcp_start")
        elif shockable == 0:
            if not step_safe("epi_now"): return
            countdown(10, T["time_left"], gif_key="rcp_start")
            shockable = ask_question("prompt_nonshock1")
            if shockable == 1:
                if not step_safe("shock_conv"): return
                countdown(10, T["time_left"], gif_key="rcp_start")
            else:
                if not step_safe("rcp_cause"): return
                countdown(10, T["time_left"], gif_key="rcp_start")
                shockable = ask_question("prompt_nonshock2")
                if shockable == 1:
                    if not step_safe("shock3"): return
                    countdown(10, T["time_left"], gif_key="rcp_start")
        rosccheck = ask_question("prompt_rosc")
        if rosccheck == 1:
            step_safe("rosc_reached")
        else:
            step_safe("rcp_loop")
        log_event(T["end_algo"])
        st.session_state.procedure_terminee = True

# Section export
if st.session_state.event_log:
    st.markdown("---")
    st.markdown(f"## {T['export']}")
    prenom = st.text_input("Prénom", key="export_prenom")
    nom = st.text_input("Nom", key="export_nom")
    if prenom and nom:
        buffer = export_log_to_excel(nom, prenom)
        filename = f"Journal_ACLS_{prenom}_{nom}.xlsx"
        st.download_button(
            label=T["download"],
            data=buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("Veuillez saisir le prénom et le nom pour activer l'export.")

run_acls()